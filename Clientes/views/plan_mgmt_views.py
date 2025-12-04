import json
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum
from ..models import Plan, Membership
from ..utils import generate_pdf_receipt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ==================== GESTION DE PLANES (ADMIN) ====================

@login_required(login_url='inicio_sesion')
def admin_plan_create(request):
    """Crear nuevo plan - Solo admin"""
    if not request.user.role or request.user.role != 'admin':
        messages.error(request, 'No tienes permisos para acceder a esta area.')
        return redirect('index_admin')
    
    if request.method == 'GET':
        # Mostrar formulario de creacion
        context = {
            'is_admin_creation': True
        }
        return render(request, 'admin_plan_create.html', context)
    
    elif request.method == 'POST':
        return process_admin_plan_creation(request)

def process_admin_plan_creation(request):
    """Procesa la creacion de plan por el admin"""
    if not request.user.role or request.user.role != 'admin':
        return JsonResponse({
            'success': False,
            'error': 'No autorizado'
        }, status=403)
    
    try:
        # Manejar tanto JSON como POST
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        # Validar campos obligatorios
        required_fields = ['name', 'plan_type', 'description', 'price', 'duration_days', 'access_days']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return JsonResponse({
                'success': False,
                'error': f'Campos requeridos faltantes: {", ".join(missing_fields)}'
            }, status=400)
        
        # CAMBIO: ELIMINADA la validación que impedía duplicar plan_type
        # (Permitimos múltiples planes del mismo tipo, ej: Premium Mensual y Premium Anual)
        
        # Crear el plan
        plan = Plan.objects.create(
            name=data['name'],
            plan_type=data['plan_type'],
            description=data['description'],
            price=data['price'],
            duration_days=data['duration_days'],
            access_days=data['access_days'],
            includes_classes=data.get('includes_classes', False) == 'true' or data.get('includes_classes') == True,
            includes_nutritionist=data.get('includes_nutritionist', False) == 'true' or data.get('includes_nutritionist') == True,
            benefits=data.get('benefits', ''),
            is_active=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Plan {plan.name} creado correctamente',
            'plan_id': plan.id
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear plan: {str(e)}'
        }, status=500)

@login_required(login_url='inicio_sesion')
def admin_plan_details(request, plan_id):
    if not request.user.role or request.user.role != 'admin':
        messages.error(request, 'No autorizado')
        return redirect('index_admin')
    
    try:
        plan = Plan.objects.get(id=plan_id)
        
        # 1. Usuarios Activos (Lista para la tabla)
        active_memberships = Membership.objects.filter(
            plan=plan, is_active=True
        ).select_related('user').order_by('end_date')

        # 2. Métricas Generales (KPIs)
        kpi_active_users = active_memberships.count()
        kpi_total_sold = Membership.objects.filter(plan=plan).count()
        
        # Ingresos Totales Históricos
        kpi_total_revenue = Membership.objects.filter(
            plan=plan, status__in=['active', 'pending']
        ).aggregate(total=Sum('amount_paid'))['total'] or 0

        # Ingresos Este Mes
        first_day_month = timezone.now().date().replace(day=1)
        kpi_monthly_revenue = Membership.objects.filter(
            plan=plan, 
            payment_date__gte=first_day_month,
            status__in=['active', 'pending']
        ).aggregate(total=Sum('amount_paid'))['total'] or 0

        # 3. Datos para el Gráfico (AÑO ACTUAL COMPLETO)
        # Reemplazamos la logica de "últimos 6 meses" por "Año Actual"
        
        # Inicializamos arrays fijos
        chart_labels = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        chart_values = [0] * 12
        
        current_year = timezone.now().year
        current_month = timezone.now().month

        # Filtramos todos los pagos de este plan en el año actual (excluyendo cancelados)
        pagos_plan_anio = Membership.objects.filter(
            plan=plan,
            payment_date__year=current_year
        ).exclude(status='cancelled')

        # Llenamos los datos mes a mes
        for pago in pagos_plan_anio:
            mes_index = pago.payment_date.month - 1
            # Usamos amount_paid para ser exactos con el dinero real ingresado
            monto = getattr(pago, 'amount_paid', plan.price) 
            chart_values[mes_index] += int(monto)

        # Recortamos para no mostrar meses futuros vacíos
        chart_labels = chart_labels[:current_month]
        chart_values = chart_values[:current_month]

        # 4. Beneficios como lista
        beneficios_lista = [b.strip() for b in plan.benefits.split(',')] if plan.benefits else []

        context = {
            'plan': plan,
            'active_memberships': active_memberships,
            
            # KPIs
            'kpi_active': kpi_active_users,
            'kpi_total_sold': kpi_total_sold,
            'kpi_total_rev': kpi_total_revenue,
            'kpi_monthly_rev': kpi_monthly_revenue,
            
            # Chart Data
            'chart_labels': json.dumps(chart_labels),
            'chart_values': json.dumps(chart_values),
            
            'beneficios_lista': beneficios_lista
        }
        
        return render(request, 'admin_plan_details.html', context)

    except Plan.DoesNotExist:
        messages.error(request, 'Plan no encontrado')
        return redirect('index_admin')

@login_required(login_url='inicio_sesion')
def admin_plan_edit(request, plan_id):
    """Editar plan - Solo admin"""
    if not request.user.role or request.user.role != 'admin':
        messages.error(request, 'No tienes permisos para acceder a esta area.')
        return redirect('index_admin')
    
    try:
        plan = Plan.objects.get(id=plan_id)
        
        if request.method == 'GET':
            # PASAR EL OBJETO 'plan' AL CONTEXTO CON EL NOMBRE CORRECTO
            context = {
                'plan': plan,  # Nombre correcto que usa el template
            }
            return render(request, 'admin_plan_edit.html', context)
        
        elif request.method == 'POST':
            # Manejar tanto JSON como POST
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST
            
            # Actualizar campos
            plan.name = data.get('name', plan.name)
            plan.plan_type = data.get('plan_type', plan.plan_type)
            plan.description = data.get('description', plan.description)
            plan.price = data.get('price', plan.price)
            plan.duration_days = data.get('duration_days', plan.duration_days)
            plan.access_days = data.get('access_days', plan.access_days)
            plan.includes_classes = data.get('includes_classes', 'False') == 'true' or data.get('includes_classes') == True
            plan.includes_nutritionist = data.get('includes_nutritionist', 'False') == 'true' or data.get('includes_nutritionist') == True
            plan.benefits = data.get('benefits', plan.benefits)
            plan.is_active = data.get('is_active', 'True') == 'true' or data.get('is_active') == True
            
            plan.save()
            
            if request.content_type == 'application/json':
                return JsonResponse({
                    'success': True,
                    'message': f'Plan {plan.name} actualizado correctamente'
                })
            else:
                messages.success(request, f'Plan {plan.name} actualizado correctamente')
                return redirect('admin_plan_details', plan_id=plan.id)
                
    except Plan.DoesNotExist:
        if request.content_type == 'application/json':
            return JsonResponse({
                'success': False,
                'error': 'Plan no encontrado'
            }, status=404)
        else:
            messages.error(request, 'Plan no encontrado')
            return redirect('index_admin')


@login_required(login_url='inicio_sesion')
def admin_plan_delete(request, plan_id):
    """Eliminar plan - Solo admin"""
    if not request.user.role or request.user.role != 'admin':
        messages.error(request, 'No tienes permisos para acceder a esta area.')
        return redirect('index_admin')
    
    try:
        plan = Plan.objects.get(id=plan_id)
        
        if request.method == 'POST':
            # Verificar si tiene membresias activas
            active_memberships = Membership.objects.filter(plan=plan, is_active=True).count()
            
            if active_memberships > 0:
                messages.error(request, f'No se puede eliminar el plan porque tiene {active_memberships} membresias activas.')
                return redirect('admin_plan_details', plan_id=plan.id)
            
            # Confirmar eliminacion
            plan_name = plan.name
            plan.delete()
            
            messages.success(request, f'Plan {plan_name} eliminado correctamente')
            return redirect('index_admin')
        
    except Plan.DoesNotExist:
        messages.error(request, 'Plan no encontrado')
        return redirect('index_admin')
    
# --- GESTIÓN DE REPORTES Y PAGOS ---

@login_required(login_url='inicio_sesion')
def exportar_pagos_excel(request):
    """Genera y descarga el reporte de pagos en Excel"""
    if not request.user.role == 'admin':
        return redirect('inicio_sesion')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_Pagos_{timezone.now().strftime("%Y%m%d")}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Transacciones"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Fondo negro
    center_align = Alignment(horizontal="center")
    
    # Encabezados
    headers = ['ID', 'Fecha', 'RUT Socio', 'Nombre Socio', 'Plan', 'Método Pago', 'Monto', 'Estado', 'Realizado Por']
    ws.append(headers)

    # Aplicar estilo a encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Obtener datos (Todos, no solo los últimos 50)
    pagos = Membership.objects.select_related('user', 'plan').all().order_by('-payment_date')

    for pago in pagos:
        # Determinar quién realizó la acción (si está en notas o log, aquí asumimos sistema o usuario actual si no hay registro)
        ws.append([
            pago.id,
            pago.payment_date.strftime("%d/%m/%Y H:i"),
            pago.user.rut,
            pago.user.get_full_name(),
            pago.plan.name,
            pago.get_payment_method_display(),
            pago.amount_paid,
            pago.get_status_display(),
            "Sistema" # Puedes personalizar esto si guardas el staff en la transacción
        ])

    # Ajustar ancho de columnas
    column_widths = [10, 20, 15, 25, 20, 15, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width

    wb.save(response)
    return response

@login_required(login_url='inicio_sesion')
def ver_recibo_pago(request, payment_id):
    """Genera el PDF del recibo de pago"""
    try:
        membership = Membership.objects.select_related('user', 'plan').get(id=payment_id)
        pdf_content = generate_pdf_receipt(membership)
        
        if pdf_content:
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="Recibo_{membership.id}.pdf"'
            return response
        else:
            messages.error(request, 'Error al generar el PDF')
            return redirect('index_admin')
            
    except Membership.DoesNotExist:
        messages.error(request, 'Transacción no encontrada')
        return redirect('index_admin')